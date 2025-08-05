from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import pandas as pd
import json
import time
import os
from datetime import datetime, timedelta

def scrape_with_selenium():
    """
    Uses Selenium to scrape USD-COP data from 30rates.com
    This method simulates a real browser and is harder to block
    """
    
    # Setup Chrome options for GitHub Actions
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Required for GitHub Actions
    chrome_options.add_argument("--no-sandbox")  # Required for GitHub Actions
    chrome_options.add_argument("--disable-dev-shm-usage")  # Required for GitHub Actions
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    driver = None
    
    try:
        print("Starting Chrome browser...")
        driver = webdriver.Chrome(options=chrome_options)
        
        print("Navigating to 30rates.com...")
        driver.get("https://30rates.com/usd-cop")
        
        # Wait for page to load
        print("Waiting for page to load...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "table"))
        )
        
        # Find the table
        print("Looking for forecast table...")
        tables = driver.find_elements(By.TAG_NAME, "table")
        
        target_table = None
        
        # Look for the table with class "tbh" or containing forecast data
        for table in tables:
            try:
                table_class = table.get_attribute("class")
                table_text = table.text
                
                if "tbh" in (table_class or "") or ("Date" in table_text and "Weekday" in table_text):
                    target_table = table
                    print("Found target table!")
                    break
            except:
                continue
        
        if not target_table:
            print("Could not find the forecast table")
            return None
        
        # Extract table data
        print("Extracting table data...")
        data = []
        
        rows = target_table.find_elements(By.TAG_NAME, "tr")
        
        for i, row in enumerate(rows):
            cells = row.find_elements(By.TAG_NAME, "td")
            
            if len(cells) >= 5:
                row_data = {
                    'Date': cells[0].text.strip(),
                    'Weekday': cells[1].text.strip(),
                    'Min': cells[2].text.strip(),
                    'Max': cells[3].text.strip(),
                    'Rate': cells[4].text.strip()
                }
                
                # Skip empty rows
                if any(row_data.values()):
                    data.append(row_data)
                    if i < 5:  # Show first 5 rows
                        print(f"Row {len(data)}: {row_data}")
        
        print(f"Successfully extracted {len(data)} rows")
        
        # Process and format the data
        formatted_data = process_and_format_data(data)
        return formatted_data
        
    except Exception as e:
        print(f"Error during scraping: {e}")
        return None
    
    finally:
        if driver:
            driver.quit()
            print("Browser closed")

def process_and_format_data(raw_data):
    """
    Process raw data to add proper date formatting, convert to numbers,
    and add weekend data (Saturday/Sunday repeating Friday values)
    """
    print("Processing and formatting data...")
    
    formatted_data = []
    current_year = datetime.now().year
    
    # Process each row from scraped data
    for i, row in enumerate(raw_data):
        try:
            # Parse the date (format: "25/07" -> "25/07/2025")
            date_parts = row['Date'].split('/')
            if len(date_parts) == 2:
                day = int(date_parts[0])
                month = int(date_parts[1])
                
                # Create full date
                full_date = datetime(current_year, month, day)
                formatted_date = full_date.strftime("%d/%m/%Y")
                
                # Convert numeric values
                min_val = float(row['Min'].replace(',', ''))
                max_val = float(row['Max'].replace(',', ''))
                rate_val = float(row['Rate'].replace(',', ''))
                
                # Add the main row (weekday)
                formatted_row = {
                    'Date': formatted_date,
                    'Weekday': row['Weekday'],
                    'Min': min_val,
                    'Max': max_val,
                    'Rate': rate_val
                }
                formatted_data.append(formatted_row)
                
                # Check if it's Friday - if so, add Saturday and Sunday with same values
                if row['Weekday'].lower() == 'friday':
                    # Add Saturday
                    saturday_date = (full_date + timedelta(days=1)).strftime("%d/%m/%Y")
                    saturday_row = {
                        'Date': saturday_date,
                        'Weekday': 'Saturday',
                        'Min': min_val,
                        'Max': max_val,
                        'Rate': rate_val
                    }
                    formatted_data.append(saturday_row)
                    
                    # Add Sunday
                    sunday_date = (full_date + timedelta(days=2)).strftime("%d/%m/%Y")
                    sunday_row = {
                        'Date': sunday_date,
                        'Weekday': 'Sunday',
                        'Min': min_val,
                        'Max': max_val,
                        'Rate': rate_val
                    }
                    formatted_data.append(sunday_row)
                    
                    print(f"Added weekend data for {formatted_date} (Friday)")
                
        except Exception as e:
            print(f"Error processing row {i}: {row} - {e}")
            continue
    
    print(f"Processed {len(raw_data)} original rows into {len(formatted_data)} total rows")
    return formatted_data

def save_data(data, output_path="."):
    """Save extracted data to Excel file - GitHub Actions version"""
    if not data:
        print("No data to save")
        return
    
    # Ensure the directory exists
    try:
        if not os.path.exists(output_path):
            os.makedirs(output_path)
            print(f"Created directory: {output_path}")
    except Exception as e:
        print(f"Warning: Could not create directory {output_path}: {e}")
        output_path = "."  # Use current directory as fallback
    
    # Create DataFrame from new data
    new_df = pd.DataFrame(data)
    
    # Generate timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Files for GitHub Actions (no historical merge in cloud)
    excel_filename = os.path.join(output_path, "forecast_usd.xlsx")
    csv_filename = os.path.join(output_path, "forecast_usd.csv")
    json_filename = os.path.join(output_path, f"forecast_usd_{timestamp}.json")
    
    try:
        # Save Excel file
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            new_df.to_excel(writer, sheet_name='USD_COP_Forecast', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['USD_COP_Forecast']
            
            # Add basic formatting
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Header formatting
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply header formatting
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Format data cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = center_alignment
                        
                        # Format numeric columns with thousands separator
                        if cell.column in [3, 4, 5]:  # Min, Max, Rate columns
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.000'
                
                # Auto-adjust column widths
                column_widths = {
                    'A': 12,  # Date
                    'B': 12,  # Weekday  
                    'C': 10,  # Min
                    'D': 10,  # Max
                    'E': 10   # Rate
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                    
            except ImportError as e:
                print(f"Warning: Could not apply advanced formatting: {e}")
        
        print(f"✅ Excel file saved: {excel_filename}")
        
        # Also save as CSV for compatibility
        new_df.to_csv(csv_filename, index=False)
        print(f"✅ CSV file saved: {csv_filename}")
        
        # Save JSON backup
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"✅ JSON backup saved: {json_filename}")
        
        print(f"📊 Total rows in files: {len(new_df)}")
        
        # Show data summary
        weekday_counts = new_df['Weekday'].value_counts()
        print(f"📈 Data breakdown: {dict(weekday_counts)}")
        
    except Exception as e:
        print(f"❌ Error saving files: {e}")
        # Fallback to just CSV
        try:
            new_df.to_csv("forecast_usd_fallback.csv", index=False)
            print(f"📁 Saved fallback CSV file")
        except Exception as e2:
            print(f"❌ Even CSV fallback failed: {e2}")
    
    # Display data preview
    print(f"\n📊 Data Preview ({len(new_df)} total rows):")
    print("=" * 60)
    print(new_df.head(10).to_string(index=False))
    
    if len(new_df) > 10:
        print(f"... and {len(new_df) - 10} more rows")
    
    print(f"\n✅ GitHub Actions process completed!")
    print(f"📈 {len(data)} new records processed")
    return excel_filename

if __name__ == "__main__":
    print("USD-COP Scraper for GitHub Actions")
    print("=" * 50)
    print("🤖 Running in GitHub Actions environment")
    print("🌐 Target: 30rates.com USD-COP forecast")
    print()
    
    # Check if required packages are installed
    try:
        import openpyxl
        print("✅ openpyxl found - Excel formatting will be applied")
    except ImportError:
        print("⚠️  openpyxl not found - basic Excel file will be created")
    
    print("\n🚀 Starting scraping process...")
    data = scrape_with_selenium()
    
    if data:
        excel_file = save_data(data)
        print(f"\n🎉 Success! {len(data)} rows extracted and saved!")
        print("📁 Files created:")
        print(f"   • Excel: forecast_usd.xlsx")
        print(f"   • CSV: forecast_usd.csv") 
        print(f"   • JSON: forecast_usd_[timestamp].json")
        print(f"\n📤 Files available as GitHub Actions artifacts")
    else:
        print("❌ Scraping failed. Check logs for details.")
        print("\n🔧 Possible issues:")
        print("1. Website structure changed")
        print("2. Network connectivity issues")
        print("3. Chrome/Selenium compatibility")
        exit(1)  # Exit with error code for GitHub Actions
